import os
import environ
import shutil
import csv
import logging
import re
from datetime import datetime
from tempfile import NamedTemporaryFile

from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.core.exceptions import PermissionDenied
from django.http import Http404, HttpRequest, HttpResponse, QueryDict
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.views import View
from django.contrib.staticfiles import finders
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font
import warnings

from urllib.parse import urlparse, parse_qs

from dojo.authorization.authorization import user_has_permission_or_403
from dojo.authorization.authorization_decorators import user_is_authorized
from dojo.authorization.roles_permissions import Permissions
from dojo.filters import (
    EndpointFilter,
    EndpointFilterWithoutObjectLookups,
    EndpointReportFilter,
    ReportFindingFilter,
    ReportFindingFilterWithoutObjectLookups,
)
from dojo.finding.queries import get_authorized_findings
from dojo.finding.views import BaseListFindings
from dojo.forms import ReportOptionsForm
from dojo.models import Dojo_User, Endpoint, Engagement, Finding, Product, Product_Type, Test,Risk_Assessment
from dojo.reports.widgets import (
    CoverPage,
    CustomReportJsonForm,
    EndpointList,
    FindingList,
    PageBreak,
    ReportOptions,
    TableOfContents,
    Widget,
    WYSIWYGContent,
    report_widget_factory,
)
from dojo.utils import (
    Product_Tab,
    add_breadcrumb,
    get_page_items,
    get_period_counts_legacy,
    get_system_setting,
    get_words_for_field,
)

logger = logging.getLogger(__name__)

EXCEL_CHAR_LIMIT = 32767


def down(request):
    return render(request, "disabled.html")


def report_url_resolver(request):
    try:
        url_resolver = request.META["HTTP_X_FORWARDED_PROTO"] + "://" + request.META["HTTP_X_FORWARDED_FOR"]
    except:
        hostname = request.META["HTTP_HOST"]
        port_index = hostname.find(":")
        if port_index != -1:
            url_resolver = request.scheme + "://" + hostname[:port_index]
        else:
            url_resolver = request.scheme + "://" + hostname
    return url_resolver + ":" + request.META["SERVER_PORT"]


class ReportBuilder(View):
    def get(self, request: HttpRequest) -> HttpResponse:
        add_breadcrumb(title="Report Builder", top_level=True, request=request)
        return render(request, self.get_template(), self.get_context(request))

    def get_findings(self, request: HttpRequest):
        findings = get_authorized_findings(Permissions.Finding_View)
        filter_string_matching = get_system_setting("filter_string_matching", False)
        filter_class = ReportFindingFilterWithoutObjectLookups if filter_string_matching else ReportFindingFilter
        return filter_class(self.request.GET, queryset=findings)

    def get_endpoints(self, request: HttpRequest):
        endpoints = Endpoint.objects.filter(finding__active=True,
                                            finding__false_p=False,
                                            finding__duplicate=False,
                                            finding__out_of_scope=False,
                                            )
        if get_system_setting("enforce_verified_status", True) or get_system_setting("enforce_verified_status_metrics", True):
            endpoints = endpoints.filter(finding__active=True)

        endpoints = endpoints.distinct()

        filter_string_matching = get_system_setting("filter_string_matching", False)
        filter_class = EndpointFilterWithoutObjectLookups if filter_string_matching else EndpointFilter
        return filter_class(request.GET, queryset=endpoints, user=request.user)

    def get_available_widgets(self, request: HttpRequest) -> list[Widget]:
        return [
            CoverPage(request=request),
            TableOfContents(request=request),
            WYSIWYGContent(request=request),
            FindingList(request=request, findings=self.get_findings(request)),
            EndpointList(request=request, endpoints=self.get_endpoints(request)),
            PageBreak()]

    def get_in_use_widgets(self, request):
        return [ReportOptions(request=request)]

    def get_template(self):
        return "dojo/report_builder.html"

    def get_context(self, request: HttpRequest) -> dict:
        return {
            "available_widgets": self.get_available_widgets(request),
            "in_use_widgets": self.get_in_use_widgets(request)}


class CustomReport(View):
    def post(self, request: HttpRequest) -> HttpResponse:
        # saving the report
        form = self.get_form(request)
        if form.is_valid():
            self._set_state(request)
            return render(request, self.get_template(), self.get_context())
        raise PermissionDenied

    def _set_state(self, request: HttpRequest):
        self.request = request
        self.host = report_url_resolver(request)
        self.selected_widgets = self.get_selected_widgets(request)
        self.widgets = list(self.selected_widgets.values())
        self.include_disclaimer = get_system_setting("disclaimer_reports_forced", 0)
        self.disclaimer = get_system_setting("disclaimer_reports")
        if self.include_disclaimer and len(self.disclaimer) == 0:
            self.disclaimer = "Please configure in System Settings."

    def get_selected_widgets(self, request):
        selected_widgets = report_widget_factory(json_data=request.POST["json"], request=request, host=self.host,
                                                      user=self.request.user, finding_notes=False, finding_images=False)

        if options := selected_widgets.get("report-options", None):
            self.report_format = options.report_type
            self.finding_notes = (options.include_finding_notes == "1")
            self.finding_images = (options.include_finding_images == "1")
        else:
            self.report_format = "HTML"
            self.finding_notes = True
            self.finding_images = True

        return report_widget_factory(json_data=request.POST["json"], request=request, host=self.host,
                              user=request.user, finding_notes=self.finding_notes,
                              finding_images=self.finding_images)

    def get_form(self, request):
        return CustomReportJsonForm(request.POST)

    def get_template(self):
        if self.report_format == "HTML":
            return "dojo/custom_html_report.html"
        raise PermissionDenied

    def get_context(self):
        return {
            "widgets": self.widgets,
            "host": self.host,
            "finding_notes": self.finding_notes,
            "finding_images": self.finding_images,
            "user_id": self.request.user.id,
            "include_disclaimer": self.include_disclaimer,
            "disclaimer": self.disclaimer,
        }


def report_findings(request):
    findings = Finding.objects.filter()
    filter_string_matching = get_system_setting("filter_string_matching", False)
    filter_class = ReportFindingFilterWithoutObjectLookups if filter_string_matching else ReportFindingFilter
    findings = filter_class(request.GET, queryset=findings)

    title_words = get_words_for_field(Finding, "title")
    component_words = get_words_for_field(Finding, "component_name")

    paged_findings = get_page_items(request, findings.qs.distinct(), 25)

    return render(request,
                  "dojo/report_findings.html",
                  {"findings": paged_findings,
                   "filtered": findings,
                   "title_words": title_words,
                    "component_words": component_words,
                   "title": "finding-list",
                   })


def report_endpoints(request):
    endpoints = Endpoint.objects.filter(finding__active=True,
                                        finding__false_p=False,
                                        finding__duplicate=False,
                                        finding__out_of_scope=False,
                                        )
    if get_system_setting("enforce_verified_status", True) or get_system_setting("enforce_verified_status_metrics", True):
        endpoints = endpoints.filter(finding__active=True)

    endpoints = endpoints.distinct()
    endpoints = EndpointFilter(request.GET, queryset=endpoints, user=request.user)

    paged_endpoints = get_page_items(request, endpoints.qs, 25)

    return render(request,
                  "dojo/report_endpoints.html",
                  {"endpoints": paged_endpoints,
                   "filtered": endpoints,
                   "title": "endpoint-list",
                   })


def report_cover_page(request):
    report_title = request.GET.get("title", "Report")
    report_subtitle = request.GET.get("subtitle", "")
    report_info = request.GET.get("info", "")

    return render(request,
                  "dojo/report_cover_page.html",
                  {"report_title": report_title,
                   "report_subtitle": report_subtitle,
                   "report_info": report_info})


@user_is_authorized(Product_Type, Permissions.Product_Type_View, "ptid")
def product_type_report(request, ptid):
    product_type = get_object_or_404(Product_Type, id=ptid)
    return generate_report(request, product_type)


@user_is_authorized(Product, Permissions.Product_View, "pid")
def product_report(request, pid):
    product = get_object_or_404(Product, id=pid)
    return generate_report(request, product)


def product_findings_report(request):
    findings = get_authorized_findings(Permissions.Finding_View)
    return generate_report(request, findings)


@user_is_authorized(Engagement, Permissions.Engagement_View, "eid")
def engagement_report(request, eid):
    engagement = get_object_or_404(Engagement, id=eid)
    return generate_report(request, engagement)


@user_is_authorized(Test, Permissions.Test_View, "tid")
def test_report(request, tid):
    test = get_object_or_404(Test, id=tid)
    return generate_report(request, test)


@user_is_authorized(Endpoint, Permissions.Endpoint_View, "eid")
def endpoint_report(request, eid):
    endpoint = get_object_or_404(Endpoint, id=eid)
    return generate_report(request, endpoint, host_view=False)


@user_is_authorized(Endpoint, Permissions.Endpoint_View, "eid")
def endpoint_host_report(request, eid):
    endpoint = get_object_or_404(Endpoint, id=eid)
    return generate_report(request, endpoint, host_view=True)


@user_is_authorized(Product, Permissions.Product_View, "pid")
def product_endpoint_report(request, pid):
    product = get_object_or_404(Product.objects.all().prefetch_related("engagement_set__test_set__test_type", "engagement_set__test_set__environment"), id=pid)
    endpoints = Endpoint.objects.filter(finding__active=True,
                                         finding__false_p=False,
                                         finding__duplicate=False,
                                         finding__out_of_scope=False)

    if get_system_setting("enforce_verified_status", True) or get_system_setting("enforce_verified_status_metrics", True):
        endpoint_ids = endpoints.filter(finding__active=True).values_list("id", flat=True)

    endpoint_ids = endpoints.values_list("id", flat=True)

    endpoints = prefetch_related_endpoints_for_report(Endpoint.objects.filter(id__in=endpoint_ids))
    endpoints = EndpointReportFilter(request.GET, queryset=endpoints)

    paged_endpoints = get_page_items(request, endpoints.qs, 25)
    report_format = request.GET.get("report_type", "HTML")
    include_finding_notes = int(request.GET.get("include_finding_notes", 0))
    include_finding_images = int(request.GET.get("include_finding_images", 0))
    include_executive_summary = int(request.GET.get("include_executive_summary", 0))
    include_table_of_contents = int(request.GET.get("include_table_of_contents", 0))
    include_disclaimer = int(request.GET.get("include_disclaimer", 0)) or (get_system_setting("disclaimer_reports_forced", 0))
    disclaimer = get_system_setting("disclaimer_reports")
    if include_disclaimer and len(disclaimer) == 0:
        disclaimer = "Please configure in System Settings."
    generate = "_generate" in request.GET
    add_breadcrumb(parent=product, title="Vulnerable Product Endpoints Report", top_level=False, request=request)
    report_form = ReportOptionsForm()
    template = "dojo/product_endpoint_pdf_report.html"

    if generate:
        report_form = ReportOptionsForm(request.GET)
        if report_format == "HTML":
            return render(request,
                          template,
                          {"product_type": None,
                           "product": product,
                           "engagement": None,
                           "test": None,
                           "endpoint": None,
                           "endpoints": endpoints.qs,
                           "findings": None,
                           "include_finding_notes": include_finding_notes,
                           "include_finding_images": include_finding_images,
                           "include_executive_summary": include_executive_summary,
                           "include_table_of_contents": include_table_of_contents,
                           "include_disclaimer": include_disclaimer,
                           "disclaimer": disclaimer,
                           "user": request.user,
                           "title": "Generate Report",
                           })
        raise Http404

    product_tab = Product_Tab(product, "Product Endpoint Report", tab="endpoints")
    return render(request,
                  "dojo/request_endpoint_report.html",
                  {"endpoints": paged_endpoints,
                   "filtered": endpoints,
                   "product_tab": product_tab,
                   "report_form": report_form,
                   "name": "Vulnerable Product Endpoints",
                   })


def generate_report(request, obj, *, host_view=False):
    user = Dojo_User.objects.get(id=request.user.id)
    product_type = None
    product = None
    engagement = None
    test = None
    endpoint = None
    endpoints = None
    report_title = None

    if type(obj).__name__ == "Product_Type":
        user_has_permission_or_403(request.user, obj, Permissions.Product_Type_View)
    elif type(obj).__name__ == "Product":
        user_has_permission_or_403(request.user, obj, Permissions.Product_View)
    elif type(obj).__name__ == "Engagement":
        user_has_permission_or_403(request.user, obj, Permissions.Engagement_View)
    elif type(obj).__name__ == "Test":
        user_has_permission_or_403(request.user, obj, Permissions.Test_View)
    elif type(obj).__name__ == "Endpoint":
        user_has_permission_or_403(request.user, obj, Permissions.Endpoint_View)
    elif type(obj).__name__ == "QuerySet" or type(obj).__name__ == "CastTaggedQuerySet" or type(obj).__name__ == "TagulousCastTaggedQuerySet":
        # authorization taken care of by only selecting findings from product user is authed to see
        pass
    else:
        if obj is None:
            msg = "No object is given to generate report for"
            raise Exception(msg)
        msg = f"Report cannot be generated for object of type {type(obj).__name__}"
        raise Exception(msg)

    report_format = request.GET.get("report_type", "HTML")
    include_finding_notes = int(request.GET.get("include_finding_notes", 0))
    include_finding_images = int(request.GET.get("include_finding_images", 0))
    include_executive_summary = int(request.GET.get("include_executive_summary", 0))
    include_table_of_contents = int(request.GET.get("include_table_of_contents", 0))
    include_disclaimer = int(request.GET.get("include_disclaimer", 0)) or (get_system_setting("disclaimer_reports_forced", 0))
    disclaimer = get_system_setting("disclaimer_reports")

    if include_disclaimer and len(disclaimer) == 0:
        disclaimer = "Please configure in System Settings."
    generate = "_generate" in request.GET
    report_name = str(obj)
    filter_string_matching = get_system_setting("filter_string_matching", False)
    report_finding_filter_class = ReportFindingFilterWithoutObjectLookups if filter_string_matching else ReportFindingFilter
    add_breadcrumb(title="Generate Report", top_level=False, request=request)
    if type(obj).__name__ == "Product_Type":
        product_type = obj
        template = "dojo/product_type_pdf_report.html"
        report_name = "Product Type Report: " + str(product_type)
        report_title = "Product Type Report"
        findings = report_finding_filter_class(request.GET, prod_type=product_type, queryset=prefetch_related_findings_for_report(Finding.objects.filter(
            test__engagement__product__prod_type=product_type)))
        products = Product.objects.filter(prod_type=product_type,
                                          engagement__test__finding__in=findings.qs).distinct()
        engagements = Engagement.objects.filter(product__prod_type=product_type,
                                                test__finding__in=findings.qs).distinct()
        tests = Test.objects.filter(engagement__product__prod_type=product_type,
                                    finding__in=findings.qs).distinct()
        if len(findings.qs) > 0:
            start_date = timezone.make_aware(datetime.combine(findings.qs.last().date, datetime.min.time()))
        else:
            start_date = timezone.now()

        end_date = timezone.now()

        r = relativedelta(end_date, start_date)
        months_between = (r.years * 12) + r.months
        # include current month
        months_between += 1

        endpoint_monthly_counts = get_period_counts_legacy(findings.qs.order_by("numerical_severity"), findings.qs.order_by("numerical_severity"), None,
                                                            months_between, start_date,
                                                            relative_delta="months")

        context = {"product_type": product_type,
                   "products": products,
                   "engagements": engagements,
                   "tests": tests,
                   "report_name": report_name,
                   "endpoint_opened_per_month": endpoint_monthly_counts[
                       "opened_per_period"] if endpoint_monthly_counts is not None else [],
                   "endpoint_active_findings": findings.qs.distinct().order_by("numerical_severity"),
                   "findings": findings.qs.distinct().order_by("numerical_severity"),
                   "include_finding_notes": include_finding_notes,
                   "include_finding_images": include_finding_images,
                   "include_executive_summary": include_executive_summary,
                   "include_table_of_contents": include_table_of_contents,
                   "include_disclaimer": include_disclaimer,
                   "disclaimer": disclaimer,
                   "user": user,
                   "team_name": settings.TEAM_NAME,
                   "title": report_title,
                   "host": report_url_resolver(request),
                   "user_id": request.user.id}

    elif type(obj).__name__ == "Product":
        product = obj
        template = "dojo/product_pdf_report.html"
        report_name = "Product Report: " + str(product)
        report_title = "Product Report"
        findings = report_finding_filter_class(request.GET, product=product, queryset=prefetch_related_findings_for_report(Finding.objects.filter(
            test__engagement__product=product)))
        ids = set(finding.id for finding in findings.qs)  # noqa: C401
        engagements = Engagement.objects.filter(test__finding__id__in=ids).distinct()
        tests = Test.objects.filter(finding__id__in=ids).distinct()
        endpoints = Endpoint.objects.filter(product=product).distinct()
        context = {"product": product,
                   "engagements": engagements,
                   "tests": tests,
                   "report_name": report_name,
                   "findings": findings.qs.distinct().order_by("numerical_severity"),
                   "include_finding_notes": include_finding_notes,
                   "include_finding_images": include_finding_images,
                   "include_executive_summary": include_executive_summary,
                   "include_table_of_contents": include_table_of_contents,
                   "include_disclaimer": include_disclaimer,
                   "disclaimer": disclaimer,
                   "user": user,
                   "team_name": settings.TEAM_NAME,
                   "title": report_title,
                   "endpoints": endpoints,
                   "host": report_url_resolver(request),
                   "user_id": request.user.id}

    elif type(obj).__name__ == "Engagement":
        logger.debug("generating report for Engagement")
        engagement = obj
        findings = report_finding_filter_class(request.GET, engagement=engagement,
                                       queryset=prefetch_related_findings_for_report(Finding.objects.filter(test__engagement=engagement)))
        report_name = "Engagement Report: " + str(engagement)
        template = "dojo/engagement_pdf_report.html"
        report_title = "Engagement Report"

        ids = set(finding.id for finding in findings.qs)  # noqa: C401
        tests = Test.objects.filter(finding__id__in=ids).distinct()
        endpoints = Endpoint.objects.filter(product=engagement.product).distinct()

        context = {"engagement": engagement,
                   "tests": tests,
                   "report_name": report_name,
                   "findings": findings.qs.distinct().order_by("numerical_severity"),
                   "include_finding_notes": include_finding_notes,
                   "include_finding_images": include_finding_images,
                   "include_executive_summary": include_executive_summary,
                   "include_table_of_contents": include_table_of_contents,
                   "include_disclaimer": include_disclaimer,
                   "disclaimer": disclaimer,
                   "user": user,
                   "team_name": settings.TEAM_NAME,
                   "title": report_title,
                   "host": report_url_resolver(request),
                   "user_id": request.user.id,
                   "endpoints": endpoints}

    elif type(obj).__name__ == "Test":
        test = obj
        findings = report_finding_filter_class(request.GET, engagement=test.engagement,
                                       queryset=prefetch_related_findings_for_report(Finding.objects.filter(test=test)))
        template = "dojo/test_pdf_report.html"
        report_name = "Test Report: " + str(test)
        report_title = "Test Report"

        context = {"test": test,
                   "report_name": report_name,
                   "findings": findings.qs.distinct().order_by("numerical_severity"),
                   "include_finding_notes": include_finding_notes,
                   "include_finding_images": include_finding_images,
                   "include_executive_summary": include_executive_summary,
                   "include_table_of_contents": include_table_of_contents,
                   "include_disclaimer": include_disclaimer,
                   "disclaimer": disclaimer,
                   "user": user,
                   "team_name": settings.TEAM_NAME,
                   "title": report_title,
                   "host": report_url_resolver(request),
                   "user_id": request.user.id}

    elif type(obj).__name__ == "Endpoint":
        endpoint = obj
        if host_view:
            report_name = "Endpoint Host Report: " + endpoint.host
            endpoints = Endpoint.objects.filter(host=endpoint.host,
                                                product=endpoint.product).distinct()
            report_title = "Endpoint Host Report"
        else:
            report_name = "Endpoint Report: " + str(endpoint)
            endpoints = Endpoint.objects.filter(pk=endpoint.id).distinct()
            report_title = "Endpoint Report"
        template = "dojo/endpoint_pdf_report.html"
        findings = report_finding_filter_class(request.GET,
                                       queryset=prefetch_related_findings_for_report(Finding.objects.filter(endpoints__in=endpoints)))

        context = {"endpoint": endpoint,
                   "endpoints": endpoints,
                   "report_name": report_name,
                   "findings": findings.qs.distinct().order_by("numerical_severity"),
                   "include_finding_notes": include_finding_notes,
                   "include_finding_images": include_finding_images,
                   "include_executive_summary": include_executive_summary,
                   "include_table_of_contents": include_table_of_contents,
                   "include_disclaimer": include_disclaimer,
                   "disclaimer": disclaimer,
                   "user": user,
                   "team_name": get_system_setting("team_name"),
                   "title": report_title,
                   "host": report_url_resolver(request),
                   "user_id": request.user.id}
    elif type(obj).__name__ in {"QuerySet", "CastTaggedQuerySet", "TagulousCastTaggedQuerySet"}:
        findings = report_finding_filter_class(request.GET, queryset=prefetch_related_findings_for_report(obj).distinct())
        report_name = "Finding"
        template = "dojo/finding_pdf_report.html"
        report_title = "Finding Report"

        context = {"findings": findings.qs.distinct().order_by("numerical_severity"),
                   "report_name": report_name,
                   "include_finding_notes": include_finding_notes,
                   "include_finding_images": include_finding_images,
                   "include_executive_summary": include_executive_summary,
                   "include_table_of_contents": include_table_of_contents,
                   "include_disclaimer": include_disclaimer,
                   "disclaimer": disclaimer,
                   "user": user,
                   "team_name": settings.TEAM_NAME,
                   "title": report_title,
                   "host": report_url_resolver(request),
                   "user_id": request.user.id}
    else:
        raise Http404

    report_form = ReportOptionsForm()

    if generate:
        report_form = ReportOptionsForm(request.GET)
        if report_format == "HTML":
            return render(request,
                          template,
                          {"product_type": product_type,
                           "product": product,
                           "engagement": engagement,
                           "report_name": report_name,
                           "test": test,
                           "endpoint": endpoint,
                           "endpoints": endpoints,
                           "findings": findings.qs.distinct().order_by("numerical_severity"),
                           "include_finding_notes": include_finding_notes,
                           "include_finding_images": include_finding_images,
                           "include_executive_summary": include_executive_summary,
                           "include_table_of_contents": include_table_of_contents,
                           "include_disclaimer": include_disclaimer,
                           "disclaimer": disclaimer,
                           "user": user,
                           "team_name": settings.TEAM_NAME,
                           "title": report_title,
                           "user_id": request.user.id,
                           "host": "",
                           "host_view": host_view,
                           "context": context,
                           })

        raise Http404
    paged_findings = get_page_items(request, findings.qs.distinct().order_by("numerical_severity"), 25)

    product_tab = None
    if engagement:
        product_tab = Product_Tab(engagement.product, title="Engagement Report", tab="engagements")
        product_tab.setEngagement(engagement)
    elif test:
        product_tab = Product_Tab(test.engagement.product, title="Test Report", tab="engagements")
        product_tab.setEngagement(test.engagement)
    elif product:
        product_tab = Product_Tab(product, title="Product Report", tab="findings")
    elif endpoints:
        if host_view:
            product_tab = Product_Tab(endpoint.product, title="Endpoint Host Report", tab="endpoints")
        else:
            product_tab = Product_Tab(endpoint.product, title="Endpoint Report", tab="endpoints")

    return render(request, "dojo/request_report.html",
                  {"product_type": product_type,
                   "product": product,
                   "product_tab": product_tab,
                   "engagement": engagement,
                   "test": test,
                   "endpoint": endpoint,
                   "findings": findings,
                   "paged_findings": paged_findings,
                   "report_form": report_form,
                   "host_view": host_view,
                   "context": context,
                   })


def prefetch_related_findings_for_report(findings):
    return findings.prefetch_related("test",
                                     "test__engagement__product",
                                     "test__engagement__product__prod_type",
                                     "risk_acceptance_set",
                                     "risk_acceptance_set__accepted_findings",
                                     "burprawrequestresponse_set",
                                     "endpoints",
                                     "tags",
                                     "notes",
                                     "files",
                                     "reporter",
                                     "mitigated_by",
                                     )


def prefetch_related_endpoints_for_report(endpoints):
    return endpoints.prefetch_related(
                                      "product",
                                      "tags",
                                     )


def get_list_index(full_list, index):
    try:
        element = full_list[index]
    except Exception:
        element = None
    return element


def get_findings(request):
    url = request.META.get("QUERY_STRING")
    if not url:
        msg = "Please use the report button when viewing findings"
        raise Http404(msg)
    url = url.removeprefix("url=")

    views = ["all", "open", "inactive", "verified",
             "closed", "accepted", "assessed", "out_of_scope",
             "false_positive", "inactive"]
    # request.path = url
    obj_name = obj_id = view = query = None
    path_items = list(filter(None, re.split(r"/|\?", url)))

    try:
        finding_index = path_items.index("finding")
    except ValueError:
        finding_index = -1
    # There is a engagement or product here
    if finding_index > 0:
        # path_items ['product', '1', 'finding', 'closed', 'test__engagement__product=1']
        obj_name = get_list_index(path_items, 0)
        obj_id = get_list_index(path_items, 1)
        view = get_list_index(path_items, 3)
        query = get_list_index(path_items, 4)
        # Try to catch a mix up
        query = query if view in views else view
    # This is findings only. Accomodate view and query
    elif finding_index == 0:
        # path_items ['finding', 'closed', 'title=blah']
        obj_name = get_list_index(path_items, 0)
        view = get_list_index(path_items, 1)
        query = get_list_index(path_items, 2)
        # Try to catch a mix up
        query = query if view in views else view
    # This is a test or engagement only
    elif finding_index == -1:
        # path_items ['test', '1', 'test__engagement__product=1']
        obj_name = get_list_index(path_items, 0)
        obj_id = get_list_index(path_items, 1)
        query = get_list_index(path_items, 2)

    filter_name = None
    if view:
        if view == "open":
            filter_name = "Open"
        elif view == "inactive":
            filter_name = "Inactive"
        elif view == "verified":
            filter_name = "Verified"
        elif view == "closed":
            filter_name = "Closed"
        elif view == "accepted":
            filter_name = "Accepted"
        elif view == "assessed":
            filter_name = "Assessed"
        elif view == "out_of_scope":
            filter_name = "Out of Scope"
        elif view == "false_positive":
            filter_name = "False Positive"

    obj = pid = eid = tid = None
    if obj_id:
        if "product" in obj_name:
            pid = obj_id
            obj = get_object_or_404(Product, id=pid)
            user_has_permission_or_403(request.user, obj, Permissions.Product_View)
        elif "engagement" in obj_name:
            eid = obj_id
            obj = get_object_or_404(Engagement, id=eid)
            user_has_permission_or_403(request.user, obj, Permissions.Engagement_View)
        elif "test" in obj_name:
            tid = obj_id
            obj = get_object_or_404(Test, id=tid)
            user_has_permission_or_403(request.user, obj, Permissions.Test_View)

    request.GET = QueryDict(query)
    list_findings = BaseListFindings(
        filter_name=filter_name,
        product_id=pid,
        engagement_id=eid,
        test_id=tid)
    findings = list_findings.get_fully_filtered_findings(request).qs

    return findings, obj


class QuickReportView(View):
    def add_findings_data(self):
        return self.findings

    def get_template(self):
        return "dojo/finding_pdf_report.html"

    def get(self, request):
        findings, obj = get_findings(request)
        self.findings = findings
        findings = self.add_findings_data()
        return self.generate_quick_report(request, findings, obj)

    def generate_quick_report(self, request, findings, obj=None):
        product = engagement = test = None

        if obj:
            if type(obj).__name__ == "Product":
                product = obj
            elif type(obj).__name__ == "Engagement":
                engagement = obj
            elif type(obj).__name__ == "Test":
                test = obj

        return render(request, self.get_template(), {
                        "report_name": "Finding Report",
                        "product": product,
                        "engagement": engagement,
                        "test": test,
                        "findings": findings,
                        "user": request.user,
                        "team_name": settings.TEAM_NAME,
                        "title": "Finding Report",
                        "user_id": request.user.id,
                  })


def get_excludes():
    return ["SEVERITIES", "age", "github_issue", "jira_issue", "objects", "risk_acceptance",
    "test__engagement__product__authorized_group", "test__engagement__product__member",
    "test__engagement__product__prod_type__authorized_group", "test__engagement__product__prod_type__member",
    "unsaved_endpoints", "unsaved_vulnerability_ids", "unsaved_files", "unsaved_request", "unsaved_response",
    "unsaved_tags", "vulnerability_ids", "cve"]


def get_foreign_keys():
    return ["defect_review_requested_by", "duplicate_finding", "finding_group", "last_reviewed_by",
        "mitigated_by", "reporter", "review_requested_by", "sonarqube_issue", "test"]


def get_attributes():
    return ["sla_age", "sla_deadline", "sla_days_remaining"]


class CSVExportView(View):
    def add_findings_data(self):
        return self.findings

    def add_extra_headers(self):
        pass

    def add_extra_values(self):
        pass

    def get(self, request):
        findings, _obj = get_findings(request)
        self.findings = findings
        findings = self.add_findings_data()
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=findings.csv"
        writer = csv.writer(response)
        allowed_attributes = get_attributes()
        excludes_list = get_excludes()
        allowed_foreign_keys = get_attributes()
        first_row = True

        for finding in findings:
            self.finding = finding
            if first_row:
                fields = []
                self.fields = fields
                for key in dir(finding):
                    try:
                        if key not in excludes_list and (not callable(getattr(finding, key)) or key in allowed_attributes) and not key.startswith("_"):
                            if callable(getattr(finding, key)) and key not in allowed_attributes:
                                continue
                            fields.append(key)
                    except Exception as exc:
                        logger.error("Error in attribute: " + str(exc))
                        fields.append(key)
                        continue
                fields.extend((
                    "test",
                    "found_by",
                    "engagement_id",
                    "engagement",
                    "product_id",
                    "product",
                    "endpoints",
                    "vulnerability_ids",
                    "tags",
                ))
                self.fields = fields
                self.add_extra_headers()

                writer.writerow(fields)

                first_row = False
            if not first_row:
                fields = []
                for key in dir(finding):
                    try:
                        if key not in excludes_list and (not callable(getattr(finding, key)) or key in allowed_attributes) and not key.startswith("_"):
                            if not callable(getattr(finding, key)):
                                value = finding.__dict__.get(key)
                            if (key in allowed_foreign_keys or key in allowed_attributes) and getattr(finding, key):
                                if callable(getattr(finding, key)):
                                    func = getattr(finding, key)
                                    result = func()
                                    value = result
                                else:
                                    value = str(getattr(finding, key))
                            if value and isinstance(value, str):
                                value = value.replace("\n", " NEWLINE ").replace("\r", "")
                            fields.append(value)
                    except Exception as exc:
                        logger.error("Error in attribute: " + str(exc))
                        fields.append("Value not supported")
                        continue
                fields.append(finding.test.title)
                fields.append(finding.test.test_type.name)
                fields.append(finding.test.engagement.id)
                fields.append(finding.test.engagement.name)
                fields.append(finding.test.engagement.product.id)
                fields.append(finding.test.engagement.product.name)

                endpoint_value = ""
                for endpoint in finding.endpoints.all():
                    endpoint_value += f"{endpoint}; "
                endpoint_value = endpoint_value.removesuffix("; ")
                if len(endpoint_value) > EXCEL_CHAR_LIMIT:
                    endpoint_value = endpoint_value[:EXCEL_CHAR_LIMIT - 3] + "..."
                fields.append(endpoint_value)

                vulnerability_ids_value = ""
                for num_vulnerability_ids, vulnerability_id in enumerate(finding.vulnerability_ids):
                    if num_vulnerability_ids > 5:
                        vulnerability_ids_value += "..."
                        break
                    vulnerability_ids_value += f"{vulnerability_id}; "
                if finding.cve and vulnerability_ids_value.find(finding.cve) < 0:
                    vulnerability_ids_value += finding.cve
                vulnerability_ids_value = vulnerability_ids_value.removesuffix("; ")
                fields.append(vulnerability_ids_value)
                # Tags
                tags_value = ""
                for num_tags, tag in enumerate(finding.tags.all()):
                    if num_tags > 5:
                        tags_value += "..."
                        break
                    tags_value += f"{tag}; "
                tags_value = tags_value.removesuffix("; ")
                fields.append(tags_value)

                self.fields = fields
                self.finding = finding
                self.add_extra_values()

                writer.writerow(fields)

        return response


class ExcelExportView(View):

    def add_findings_data(self):
        return self.findings

    def add_extra_headers(self):
        pass

    def add_extra_values(self):
        pass

    def get(self, request):
        findings, _obj = get_findings(request)
        self.findings = findings
        findings = self.add_findings_data()
        workbook = Workbook()
        workbook.iso_dates = True
        worksheet = workbook.active
        worksheet.title = "Findings"
        self.worksheet = worksheet
        font_bold = Font(bold=True)
        self.font_bold = font_bold
        allowed_attributes = get_attributes()
        excludes_list = get_excludes()
        allowed_foreign_keys = get_attributes()

        row_num = 1
        for finding in findings:
            if row_num == 1:
                col_num = 1
                for key in dir(finding):
                    try:
                        if key not in excludes_list and (not callable(getattr(finding, key)) or key in allowed_attributes) and not key.startswith("_"):
                            if callable(getattr(finding, key)) and key not in allowed_attributes:
                                continue
                            cell = worksheet.cell(row=row_num, column=col_num, value=key)
                            cell.font = font_bold
                            col_num += 1
                    except Exception as exc:
                        logger.error("Error in attribute: " + str(exc))
                        cell = worksheet.cell(row=row_num, column=col_num, value=key)
                        col_num += 1
                        continue
                cell = worksheet.cell(row=row_num, column=col_num, value="found_by")
                cell.font = font_bold
                col_num += 1
                worksheet.cell(row=row_num, column=col_num, value="engagement_id")
                cell = cell.font = font_bold
                col_num += 1
                cell = worksheet.cell(row=row_num, column=col_num, value="engagement")
                cell.font = font_bold
                col_num += 1
                cell = worksheet.cell(row=row_num, column=col_num, value="product_id")
                cell.font = font_bold
                col_num += 1
                cell = worksheet.cell(row=row_num, column=col_num, value="product")
                cell.font = font_bold
                col_num += 1
                cell = worksheet.cell(row=row_num, column=col_num, value="endpoints")
                cell.font = font_bold
                col_num += 1
                cell = worksheet.cell(row=row_num, column=col_num, value="vulnerability_ids")
                cell.font = font_bold
                col_num += 1
                cell = worksheet.cell(row=row_num, column=col_num, value="tags")
                cell.font = font_bold
                col_num += 1
                self.row_num = row_num
                self.col_num = col_num
                self.add_extra_headers()

                row_num = 2
            if row_num > 1:
                col_num = 1
                for key in dir(finding):
                    try:
                        if key not in excludes_list and (not callable(getattr(finding, key)) or key in allowed_attributes) and not key.startswith("_"):
                            if not callable(getattr(finding, key)):
                                value = finding.__dict__.get(key)
                            if (key in allowed_foreign_keys or key in allowed_attributes) and getattr(finding, key):
                                if callable(getattr(finding, key)):
                                    func = getattr(finding, key)
                                    result = func()
                                    value = result
                                else:
                                    value = str(getattr(finding, key))
                            if value and isinstance(value, datetime):
                                value = value.replace(tzinfo=None)
                            worksheet.cell(row=row_num, column=col_num, value=value)
                            col_num += 1
                    except Exception as exc:
                        logger.error("Error in attribute: " + str(exc))
                        worksheet.cell(row=row_num, column=col_num, value="Value not supported")
                        col_num += 1
                        continue
                worksheet.cell(row=row_num, column=col_num, value=finding.test.test_type.name)
                col_num += 1
                worksheet.cell(row=row_num, column=col_num, value=finding.test.engagement.id)
                col_num += 1
                worksheet.cell(row=row_num, column=col_num, value=finding.test.engagement.name)
                col_num += 1
                worksheet.cell(row=row_num, column=col_num, value=finding.test.engagement.product.id)
                col_num += 1
                worksheet.cell(row=row_num, column=col_num, value=finding.test.engagement.product.name)
                col_num += 1

                endpoint_value = ""
                for endpoint in finding.endpoints.all():
                    endpoint_value += f"{endpoint}; \n"
                endpoint_value = endpoint_value.removesuffix("; \n")
                if len(endpoint_value) > EXCEL_CHAR_LIMIT:
                    endpoint_value = endpoint_value[:EXCEL_CHAR_LIMIT - 3] + "..."
                worksheet.cell(row=row_num, column=col_num, value=endpoint_value)
                col_num += 1

                vulnerability_ids_value = ""
                for num_vulnerability_ids, vulnerability_id in enumerate(finding.vulnerability_ids):
                    if num_vulnerability_ids > 5:
                        vulnerability_ids_value += "..."
                        break
                    vulnerability_ids_value += f"{vulnerability_id}; \n"
                if finding.cve and vulnerability_ids_value.find(finding.cve) < 0:
                    vulnerability_ids_value += finding.cve
                vulnerability_ids_value = vulnerability_ids_value.removesuffix("; \n")
                worksheet.cell(row=row_num, column=col_num, value=vulnerability_ids_value)
                col_num += 1
                # tags
                tags_value = ""
                for tag in finding.tags.all():
                    tags_value += f"{tag}; \n"
                tags_value = tags_value.removesuffix("; \n")
                worksheet.cell(row=row_num, column=col_num, value=tags_value)
                col_num += 1
                self.col_num = col_num
                self.row_num = row_num
                self.finding = finding
                self.add_extra_values()
            row_num += 1

        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(
            content=stream,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=findings.xlsx"
        return response


class PSRAExcelExportView(View):
    factor_columns = {
        "EE": "E",  # Ease of Exploit
        "ED": "F",  # Ease of Discovery
        "AW": "G",  # Awareness
        "DT": "H",  # Detectability
        "SR": "S",  # Security Researcher
        "AT": "T",  # Advanced Network Threat
        "OS": "U",  # Outsider
        "HD": "V",  # Hardware Defects
        "SF": "W",  # Software Defects
        "IN": "X",  # Intruder
        "MC": "Y",  # Malicious Code
        "IO": "Z",  # Infrastructure Outage
        "IS": "AA",  # Insider
        "TI": "AB",  # Trusted Insider
        "CU": "AC",  # Clinical Users
        "SA": "AD",  # System Admins
        "ND": "AE",  # Natural or Man-Made Disaster
        "EN": "AF",  # Engineer
        "AA": "AG",  # Automated or Remote Access
        "SD": "AI",  # Sensitive Data
        "PD": "AJ",  # Personal Data
        "HN": "AK",  # Hospital Network
        "ATD": "AL",  # Audit Trail Data
        "CF": "AM",  # Configuration
        "SS": "AN",  # System Software
        "HW": "AO",  # Hardware
        "RE": "AP",  # Removable Media with ePHI
        "RM": "AQ",  # Removable Media
        "LD": "AR",  # Logging Data
        "PDN": "AS",  # Product Documentation
        "PS": "AT",  # Personal
        "PR": "AU",  # Product
        "NW": "AV",  # Network
        "AD": "AW",  # All Data
        "CI": "AX",  # Confidentiality
        "IT": "AY",  # Integrity
        "AV": "AZ",  # Availability
        "DL": "BK",  # Delta Likelihood
        "DV": "BL",  # Delta Vulnerability
        "DC": "BN",  # Delta Confidentiality
        "DI": "BO",  # Delta Integrity
        "DA": "BP",  # Delta Availability
        "BL": "CA",  # Business Likelihood
        "FD": "BU",  # Financial Damage
        "RD": "BV",  # Reputation Damage
        "RN": "BW",  # Regulatory Non-Compliance
        "CN": "BX",  # Customer Non-Compliance
        "PV": "BY",  # Privacy Violation
    }

    def get_product(self, product_id: int):
        return get_object_or_404(Product, id=product_id)

    def get_test(self, test_id: int):
        return get_object_or_404(Test, id=test_id)
    

    def initialize_factors_with_dash(self, worksheet, row):
        """Initialize cells corresponding to factors in factors_to_mark_with_x with '-'."""
        for factor in self.factors_to_mark_with_x:
            column_letter = self.factor_columns.get(factor)
            if column_letter:
                cell_ref = f"{column_letter}{row}"
                worksheet[cell_ref].value = '-'

    
    def add_findings_data(self):
        return self.findings
    
    # List of factors that should be marked with 'x' if present
    factors_to_mark_with_x = {
        "SR", "AT", "OS", "HD", "SF", "IN", "MC", "IO", "IS", "TI", "CU", "SA", 
        "ND", "EN", "AA", "SD", "PD", "HN", "ATD", "CF", "SS", "HW", "RE", "RM", 
        "LD", "PDN", "PS", "PR", "NW", "AD"
    }

    def parse_vector_string(self, vector_string):
        """Parse a vector string into a dictionary of factor-value pairs, ignoring case for factor codes."""
        vector_string = vector_string.replace("PSRA:1.0 ", "", 1).strip()
        factor_values = {}
        
        for pair in vector_string.split("/"):
            if ':' in pair:
                factor, value = map(str.strip, pair.split(':', 1))
                factor_values[factor.upper()] = value.strip()
        
        return factor_values

    def fill_finding_row(self, worksheet, factor_values, row):
        for factor_code, value in factor_values.items():
            column_letter = self.factor_columns.get(factor_code.upper())
            if column_letter:
                cell_ref = f"{column_letter}{row}"
                
                # If factor code is in `factors_as_x`, set value as 'x'
                if factor_code.upper() in self.factors_to_mark_with_x:
                    worksheet[cell_ref].value = 'X'
                else:
                    # Strip special characters and whitespace from the value
                    clean_value = value.strip().strip("'")  # Removes whitespace and single quotes
                    
                    # Convert to integer if it's a numeric string; otherwise, keep it as a string
                    try:
                        worksheet[cell_ref].value = int(clean_value) if clean_value.isdigit() else clean_value
                    except ValueError:
                        worksheet[cell_ref].value = clean_value  # Assign directly if non-numeric
    
    def set_threat_type(self,rmm_worksheet, idx, threat_types):
        # Mapping individual threat types to their respective column indexes
        threat_mapping = {
            'S': 'M',  # Spoofing
            'T': 'N',  # Tampering
            'R': 'O',  # Repudiation
            'I': 'P',  # Information Disclosure
            'D': 'Q',  # Denial of Service
            'E': 'R'   # Elevation of Privilege
        }

        # First, set all cells in M to R to '-'
        for col in threat_mapping.values():
            rmm_worksheet[f"{col}{idx}"] = '-'
        
        # Split the comma-separated threat types and set the respective cells to 'X'
        for threat in threat_types:
            threat = threat.strip()  # Ensure there are no spaces
            if threat in threat_mapping:
                rmm_worksheet[f"{threat_mapping[threat]}{idx}"] = 'X'


    def get(self, request, test_id=None):
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.worksheet._reader")
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.reader.workbook")

        if test_id:
            test = self.get_test(test_id)
            product = test.engagement.product

        else:
            raw_url = request.GET.get("url", "")
            product_id = request.GET.get("product_id")

            try:
                url_parts = raw_url.split('?')
                base_path = url_parts[0]

                if not product_id:
                    match = re.search(r'/product/(\d+)/', base_path)
                    if match:
                        product_id = match.group(1)

                if product_id:
                    product_id = int(product_id)
                else:
                    return HttpResponse("Error: Cannot determine product ID", status=400)

                product = self.get_product(product_id)

                if len(url_parts) > 1:
                    additional_params = parse_qs(url_parts[1])

                    get_copy = request.GET.copy()
                    for key, value in additional_params.items():
                        get_copy[key] = value[0] if len(value) == 1 else value

                    request.GET = get_copy

            except (ValueError, TypeError) as e:
                logger.error(f"Error parsing URL: {raw_url}. Error: {str(e)}")
                return HttpResponse(f"Invalid URL: {raw_url}", status=400)

        pid = product.id
        root_path = environ.Path(__file__) - 3  # Three folders back 
        product_folder = f"product_{pid}"
        base_upload_dir = root_path('uploads', 'scan_reports', product_folder)

        os.makedirs(base_upload_dir, exist_ok=True)
        existing_file_path = os.path.join(base_upload_dir, f"RiskAssessment_{pid}.xlsx")

        if not os.path.exists(existing_file_path):
            template_path = root_path('uploads', 'PSRA/PSRA template.xlsx')
            # template_path = f"{settings.STATIC_ROOT}/PSRA/PSRA template.xlsx"
            shutil.copy(template_path, existing_file_path)

        workbook = load_workbook(existing_file_path, data_only=False)
        rmm_worksheet = workbook["RMM"]
        mitigation_worksheet = workbook["Mitigations"]

        findings, _obj = get_findings(request)
        self.findings = findings
        findings = self.add_findings_data()

        start_row = 4
        for idx, finding in enumerate(findings, start=start_row):
            risk_assessment = Risk_Assessment.objects.filter(assessed_findings=finding).first()
            if risk_assessment is None: 
                continue
            
            vector_string = risk_assessment.vector_string
            if not vector_string:
                continue
            
            vulnerability_description = finding.description
            threat_description = risk_assessment.threat_description
            risk_statement = risk_assessment.risk_statement
            hazard_statement = risk_assessment.related_hazard_item
            vulnerability_cause = risk_assessment.vulnerability_cause
            threat_types = risk_assessment.threat_types
            rationale_and_actions = risk_assessment.rational_and_actions
            accepted_risk = finding.risk_accepted
            mitigation_type = risk_assessment.mitigation_types
            mitigation_reference = risk_assessment.mitigation_reference
            
            self.set_threat_type(rmm_worksheet, idx, threat_types)
            self.initialize_factors_with_dash(rmm_worksheet, idx)
            factor_values = self.parse_vector_string(vector_string)
            self.fill_finding_row(rmm_worksheet, factor_values, idx)

            rmm_worksheet[f"B{idx}"] = "V"+str(finding.id)
            rmm_worksheet[f"C{idx}"] = vulnerability_description
            rmm_worksheet[f"K{idx}"] = threat_description
            rmm_worksheet[f"L{idx}"] = risk_statement
            rmm_worksheet[f"CF{idx}"] = hazard_statement
            rmm_worksheet[f"D{idx}"] = vulnerability_cause
            rmm_worksheet[f"CE{idx}"] = rationale_and_actions
            rmm_worksheet[f"CD{idx}"] = "Yes" if accepted_risk else "No"

            mitigation_worksheet[f"A{idx-2}"] = "M"+str(finding.id)
            mitigation_worksheet[f"B{idx-2}"] = finding.mitigation
            mitigation_worksheet[f"C{idx-2}"] = mitigation_type
            mitigation_worksheet[f"D{idx-2}"] = mitigation_reference
            rmm_worksheet[f"BI{idx}"] = "M"+str(finding.id)

        workbook.save(existing_file_path)

        with open(existing_file_path, 'rb') as file:
            response = HttpResponse(
                file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = f'attachment; filename="RiskAssessment_{pid}.xlsx"'
        
        return response


